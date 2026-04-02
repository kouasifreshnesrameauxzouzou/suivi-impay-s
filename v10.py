"""
NSIA Vie Assurances — Portail Impayés (mode LOCAL)
===================================================
Architecture :
  1. Extraction initiale 2015→aujourd'hui par semestres (chunked)
     → stockée dans  NSIA_Impayes_BASE.parquet  (rapide, compact)
  2. L'app charge le Parquet au démarrage (< 2 s même pour 1M lignes)
  3. Tous les filtres s'appliquent en mémoire → instantané
  4. Bouton "🔄 Mettre à jour" : extrait depuis la dernière date + 1j,
     fusionne et réécrit le Parquet

Règles métier figées :
  ✔ INDIVIDUEL uniquement (JAQUITP → JAPOLIP → JAIDENP)
  ✔ INENC = ' '
  ✔ Polices commençant par 8 → exclues
  ✔ Produits {5100 5200 5300 6100 6120 6400 6420 6625 7520 7525 7550} → exclus
"""

import streamlit as st
import pandas as pd
import numpy as np
import pyodbc
import io
import json
import os
import traceback
from datetime import datetime, date, timedelta
from dateutil.relativedelta import relativedelta

# ══════════════════════════════════════════════════════════════════════════════
#  CHEMINS FICHIERS LOCAUX
# ══════════════════════════════════════════════════════════════════════════════
BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
DATA_FILE   = os.path.join(BASE_DIR, "NSIA_Impayes_BASE.parquet")
META_FILE   = os.path.join(BASE_DIR, "NSIA_Impayes_META.json")
DATE_DEBUT_INIT = "20150101"   # première extraction depuis 2015

# ══════════════════════════════════════════════════════════════════════════════
#  CONSTANTES MÉTIER
# ══════════════════════════════════════════════════════════════════════════════
NAVY = "#0B2559"
GOLD = "#C8A951"
GRAY = "#6b7280"
RED  = "#b91c1c"

PRODUITS_EXCLUS = {5100,5200,5300,6100,6120,6400,6420,6625,7520,7525,7550}
PROD_EXCL_SQL   = ", ".join(str(p) for p in sorted(PRODUITS_EXCLUS))

DB = dict(
    DRIVER   = "{ODBC Driver 17 for SQL Server}",
    SERVER   = r"10.8.3.9\SUNCOTEDIVOIRE",
    DATABASE = "SUN_COTEDIVOIRE",
    UID      = "reportdata",
    PWD      = "reportdata$2025",
    Timeout  = "600",   # 10 min pour les gros chunks
)

MODE_MAP = {
    "B":"Bancaire","C":"Chèque","E":"Espèce","F":"Réemploi (Rente)",
    "L":"Mobile / Numérique","M":"Mandat","R":"Réemploi",
    "T":"Transfert compte courant","V":"Virement bancaire","X":"Abandon de créances",
}
MOTIF_MAP = {
    "01":"Pas d'autorisation prélèvement","02":"Compte en surveillance",
    "03":"Compte bloqué","04":"Compte clos","05":"Compte inexistant",
    "06":"Provision insuffisante","08":"Clé RIB erronée","10":"Contentieux",
    "12":"En attente d'autorisation","88":"Défaut de provision",
    "99":"Ecobank compte inexistant",
}
PACCO_MAP = {
    "A":"Annuelle","S":"Semestrielle","T":"Trimestrielle","M":"Mensuelle","U":"Unique",
}

# ══════════════════════════════════════════════════════════════════════════════
#  PAGE CONFIG + CSS
# ══════════════════════════════════════════════════════════════════════════════
st.set_page_config(page_title="💸 Impayés · NSIA Vie", page_icon="💸", layout="wide")

st.markdown(f"""
<style>
@import url('https://fonts.googleapis.com/css2?family=Sora:wght@300;400;600;700&display=swap');
html,body,[class*="css"]{{font-family:'Sora',sans-serif;}}
#MainMenu,footer,header{{visibility:hidden;}}
.block-container{{padding:1.2rem 1.8rem!important;max-width:100%!important;}}

.header-banner{{background:linear-gradient(135deg,{NAVY} 0%,#1B3A6B 100%);
  color:white;padding:1.3rem 2rem;border-radius:12px;margin-bottom:1.2rem;
  display:flex;align-items:center;justify-content:space-between;
  box-shadow:0 4px 20px rgba(11,37,89,.35);}}
.header-banner h1{{margin:0;font-size:1.45rem;font-weight:700;letter-spacing:-.3px;}}
.header-banner p{{margin:.2rem 0 0;font-size:.78rem;opacity:.7;}}
.header-badge{{background:{GOLD};color:{NAVY};font-weight:700;
  font-size:.72rem;padding:.3rem .9rem;border-radius:20px;white-space:nowrap;}}

.db-bar{{background:#f8f9fc;border:1px solid #e2e6f0;border-radius:10px;
  padding:.8rem 1.2rem;margin-bottom:1rem;display:flex;align-items:center;
  gap:1.2rem;flex-wrap:wrap;}}
.db-info{{font-size:.78rem;color:{NAVY};font-weight:600;}}
.db-sub{{font-size:.7rem;color:{GRAY};margin-top:2px;}}
.db-ok{{color:#16a34a;font-weight:700;}}
.db-warn{{color:#d97706;font-weight:700;}}

.kpi-grid{{display:grid;grid-template-columns:repeat(5,1fr);gap:.8rem;margin-bottom:1.2rem;}}
.kpi-card{{background:white;border:1px solid #E8EBF3;border-radius:10px;
  padding:1rem 1.2rem;box-shadow:0 2px 8px rgba(11,37,89,.06);
  position:relative;overflow:hidden;}}
.kpi-card::before{{content:'';position:absolute;top:0;left:0;width:4px;height:100%;background:{NAVY};}}
.kpi-card.gold::before{{background:{GOLD};}}
.kpi-card.red::before{{background:{RED};}}
.kpi-card.green::before{{background:#38A169;}}
.kpi-label{{font-size:.68rem;color:{GRAY};text-transform:uppercase;letter-spacing:.6px;font-weight:600;}}
.kpi-value{{font-size:1.5rem;font-weight:700;color:{NAVY};margin:.2rem 0 0;line-height:1.1;}}
.kpi-value.red{{color:{RED};}}
.kpi-sub{{font-size:.7rem;color:#A0AEC0;margin-top:.15rem;}}

.section-title{{font-size:.82rem;font-weight:700;color:{NAVY};text-transform:uppercase;
  letter-spacing:.8px;border-bottom:2px solid {GOLD};padding-bottom:.35rem;margin:1rem 0 .7rem;}}
.fsec{{font-size:11px;font-weight:700;color:{NAVY};text-transform:uppercase;
  letter-spacing:.5px;border-left:3px solid {GOLD};padding-left:8px;margin:10px 0 6px;}}

[data-testid="stSidebar"]{{background:#F7F8FC;}}
.stButton>button{{background:{NAVY}!important;color:white!important;border:none!important;
  border-radius:8px!important;font-weight:600!important;font-family:'Sora',sans-serif!important;}}
.stButton>button:hover{{background:#1B3A6B!important;}}
.stDownloadButton>button{{background:{GOLD}!important;color:{NAVY}!important;
  border:none!important;border-radius:8px!important;font-weight:700!important;}}
.stDataFrame{{border-radius:8px;overflow:hidden;}}
label[data-testid="stWidgetLabel"] p{{font-size:11px!important;font-weight:600!important;color:{NAVY}!important;}}
</style>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
#  MÉTADONNÉES LOCALES
# ══════════════════════════════════════════════════════════════════════════════
def load_meta() -> dict:
    if os.path.exists(META_FILE):
        with open(META_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}

def save_meta(meta: dict):
    with open(META_FILE, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)

def base_exists() -> bool:
    return os.path.exists(DATA_FILE) and os.path.getsize(DATA_FILE) > 0

# ══════════════════════════════════════════════════════════════════════════════
#  CONNEXION SQL SERVER
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_resource(show_spinner=False)
def get_conn():
    cs = ";".join(f"{k}={v}" for k, v in DB.items())
    return pyodbc.connect(cs, autocommit=True)

def new_conn():
    """Connexion fraîche (non cachée) — utilisée pour les extractions longues."""
    cs = ";".join(f"{k}={v}" for k, v in DB.items())
    return pyodbc.connect(cs, autocommit=True)

# ══════════════════════════════════════════════════════════════════════════════
#  SQL PAR CHUNK (DDPCO entre d1 et d2)
# ══════════════════════════════════════════════════════════════════════════════
def build_sql_chunk(d1: str, d2: str) -> str:
    return f"""
    SELECT
        Q.WNUCO                                                             AS [NUMERO QUITTANCE],
        Q.WNUPO                                                             AS [NUMERO POLICE],
        Q.ANUCO                                                             AS [POLICE EXTERNE],
        Q.WNPRO                                                             AS [CODE PRODUIT],
        P.JAPOLIP_JAASSUP_WNUAD                                            AS [NUMERO ASSURE],
        Q.WINAG                                                             AS [CODE AGENT],
        A.JAAGENP_NOMTOT                                                   AS [NOM AGENT],
        Q.WUCLI                                                             AS [CODE CLIENT],
        Q.TENCO                                                             AS [MODE ENCAISSEMENT],
        Q.PACCO                                                             AS [PERIODICITE],
        Q.TOTCO                                                             AS [MONTANT QUITTANCE],
        Q.MPYCO                                                             AS [MONTANT A PAYER],
        Q.SOLPCO                                                            AS [SOLDE POLICE],
        Q.SOLDCO                                                            AS [SOLDE ENCAISSEMENT],
        TRY_CONVERT(DATE, CONVERT(VARCHAR(8), Q.DDQCO), 112)               AS [DATE QUITTANCE],
        TRY_CONVERT(DATE, CONVERT(VARCHAR(8), Q.DDPCO), 112)               AS [DEBUT PERIODE],
        TRY_CONVERT(DATE, CONVERT(VARCHAR(8), Q.DFECO), 112)               AS [FIN PERIODE],
        TRY_CONVERT(DATE, CONVERT(VARCHAR(8), Q.DCPCO), 112)               AS [DATE COMPTABLE],
        TRY_CONVERT(DATE, CONVERT(VARCHAR(8), Q.DCRAT_), 112)              AS [DATE CREATION],
        TRY_CONVERT(DATE, CONVERT(VARCHAR(8), P.JAPOLIP_DEFPO), 112)       AS [DATE EFFET],
        TRY_CONVERT(DATE, CONVERT(VARCHAR(8), P.JAPOLIP_DFEPO), 112)       AS [DATE FIN EFFET],
        TRY_CONVERT(DATE, CONVERT(VARCHAR(8), P.JAPOLIP_DRSPO), 112)       AS [DATE RESILIATION],
        P.JAPOLIP_MRGPO                                                    AS [MODE REGLEMENT],
        I.JAIDENP_TITAD                                                    AS [CIVILITE],
        I.JAIDENP_NOMAD                                                    AS [NOM],
        I.JAIDENP_PREAD                                                    AS [PRENOMS],
        I.JAIDENP_NOMTOT                                                   AS [NOM COMPLET],
        I.JAIDENP_TELAD                                                    AS [TELEPHONE],
        TRY_CAST(I.FICXML AS XML).value('(//*[local-name()="TEL2D"])[1]','NVARCHAR(50)')   AS [TELEPHONE 2],
        TRY_CONVERT(DATE, CONVERT(VARCHAR(8), I.JAIDENP_DNAAD), 112)      AS [DATE NAISSANCE],
        TRY_CAST(I.FICXML AS XML).value('(//*[local-name()="ADEAD"])[1]','NVARCHAR(255)') AS [EMAIL],
        dbo.adresse_id(I.JAIDENP_WNUAD)                                   AS [ADRESSE POSTALE],
        -- Motif rejet (JASPRLP)
        PR.JASPRLP_MRFPV                                                   AS [MOTIF PRELEVEMENT],
        -- Informations bancaires (JAENCAP + JAPBENP)
        TRY_CONVERT(DATE, CONVERT(VARCHAR(8), JE.DCRAT_), 112)             AS [DATE AFFECTATION],
        JE.RFBGT                                                           AS [REF BANCAIRE ENC],
        JB.REFBQ                                                           AS [REF BANCAIRE BANQUE],
        JB.LBABQ                                                           AS [LIBELLE BANQUE],
        JB.LAGBQ                                                           AS [LIBELLE AGENCE],
        JB.IBANQ                                                           AS [CODE IBAN],
        JB.LIBEN                                                           AS [LIBELLE ENCAISSEMENT DIRECT]
    FROM NSIACIF.JAQUITP Q WITH (NOLOCK)
    INNER JOIN NSIACIF.JAPOLIP P WITH (NOLOCK)
        ON P.JAPOLIP_WNUPO = Q.WNUPO
    INNER JOIN NSIACIF.JAIDENP I WITH (NOLOCK)
        ON I.JAIDENP_WNUAD = P.JAPOLIP_JAASSUP_WNUAD
    LEFT JOIN NSIACIF.JAAGENP A WITH (NOLOCK)
        ON A.JAAGENP_WINAG = P.JAPOLIP_JASCCDP_WINAG
    LEFT JOIN (
        SELECT JASPRLP_JAQUITP_WNUCO, JASPRLP_MRFPV,
               ROW_NUMBER() OVER (PARTITION BY JASPRLP_JAQUITP_WNUCO ORDER BY DCRAT_ DESC) AS RN
        FROM NSIACIF.JASPRLP WITH (NOLOCK)
    ) PR ON PR.JASPRLP_JAQUITP_WNUCO = Q.WNUCO AND PR.RN = 1
    -- Informations bancaires : JAENCAP (émissions individuelles) → JAPBENP (référentiel banque)
    LEFT JOIN NSIACIF.JAENCAP JE WITH (NOLOCK)
        ON JE.WNUCO = Q.WNUCO
    LEFT JOIN NSIACIP.JAPBENP JB WITH (NOLOCK)
        ON JB.REFBQ = JE.RFBGT
    WHERE Q.DDPCO BETWEEN '{d1}' AND '{d2}'
      AND Q.MPYCO > 0
      AND Q.INENC = ' '
      AND TRY_CAST(Q.WNPRO AS INT) NOT IN ({PROD_EXCL_SQL})
      AND Q.WNUPO NOT LIKE '8%'
    """

def fetch_chunk(d1: str, d2: str) -> pd.DataFrame:
    """Connexion fraîche par chunk pour éviter les déconnexions IMC06."""
    conn = new_conn()
    try:
        cursor = conn.cursor()
        cursor.execute(build_sql_chunk(d1, d2))
        cols = [d[0] for d in cursor.description]
        rows = cursor.fetchall()
        cursor.close()
        return pd.DataFrame.from_records(rows, columns=cols)
    finally:
        try:
            conn.close()
        except Exception:
            pass

def generer_semestres(d_debut: str, d_fin: str) -> list[tuple[str,str]]:
    """Découpe la plage en semestres pour éviter les timeouts."""
    start = datetime.strptime(d_debut, "%Y%m%d")
    end   = datetime.strptime(d_fin,   "%Y%m%d")
    chunks = []
    cur = start
    while cur <= end:
        chunk_end = min(cur + relativedelta(months=6) - timedelta(days=1), end)
        chunks.append((cur.strftime("%Y%m%d"), chunk_end.strftime("%Y%m%d")))
        cur = chunk_end + timedelta(days=1)
    return chunks

# ══════════════════════════════════════════════════════════════════════════════
#  ENRICHISSEMENT (colonnes analytiques, appliqué après chargement)
# ══════════════════════════════════════════════════════════════════════════════
def enrich(df: pd.DataFrame) -> pd.DataFrame:
    if df.empty:
        return df
    today = pd.Timestamp.today().normalize()

    # Dates
    DATE_COLS = ["DATE QUITTANCE","DEBUT PERIODE","FIN PERIODE","DATE COMPTABLE",
                 "DATE CREATION","DATE EFFET","DATE FIN EFFET","DATE RESILIATION",
                 "DATE NAISSANCE","DATE AFFECTATION"]
    for c in DATE_COLS:
        if c in df.columns:
            df[c] = pd.to_datetime(df[c], errors="coerce")

    df["MONTANT A PAYER"] = pd.to_numeric(df["MONTANT A PAYER"], errors="coerce").fillna(0)

    df["LIBELLE MODE"]            = df["MODE ENCAISSEMENT"].astype(str).str.strip().map(MODE_MAP).fillna("Non défini")
    df["LIBELLE MOTIF"]           = df["MOTIF PRELEVEMENT"].astype(str).str.strip().map(MOTIF_MAP).fillna("Sans motif")
    df["LIBELLE PERIODICITE"]     = df["PERIODICITE"].astype(str).str.strip().map(PACCO_MAP).fillna(df["PERIODICITE"])
    # Mode règlement police (JAPOLIP_MRGPO) — même mapping que TENCO
    if "MODE REGLEMENT" in df.columns:
        df["LIBELLE MODE REGLEMENT"] = df["MODE REGLEMENT"].astype(str).str.strip().map(MODE_MAP).fillna("Non défini")

    df["DUREE IMPAYES"] = ((today - df["DEBUT PERIODE"]).dt.days / 365.25).round(2)
    df["CLASSE IMPAYES"] = pd.cut(
        df["DUREE IMPAYES"], bins=[-np.inf,1,2,3,np.inf],
        labels=["< 1 an","1-2 ans","2-3 ans","> 3 ans"], right=False,
    ).astype(str).replace("nan","Indéfini")

    if "DATE EFFET" in df.columns:
        df["ANCIENNETE ANS"] = ((today - df["DATE EFFET"]).dt.days / 365.25).round(1)
        df["CLASSE ANCIENNETE"] = pd.cut(
            df["ANCIENNETE ANS"], bins=[-np.inf,1,2,3,5,10,np.inf],
            labels=["≤ 1 an","1-2 ans","2-3 ans","3-5 ans","5-10 ans","> 10 ans"], right=True,
        ).astype(str).replace("nan","Indéfini")

    resil = pd.to_datetime(df.get("DATE RESILIATION"), errors="coerce")
    fin   = pd.to_datetime(df.get("DATE FIN EFFET"),   errors="coerce")
    df["STATUT POLICE"] = np.where(
        resil.notna(), "Résiliée",
        np.where(fin.notna() & (fin < today), "Expirée", "Active"),
    )

    civ = df["CIVILITE"].fillna("").astype(str).str.upper().str.strip()
    df["GENRE"] = np.where(
        civ.isin({"M.","M","MLE","MR"}), "Homme",
        np.where(civ.isin({"MME","MLLE","F"}), "Femme", "Non défini"),
    )

    if "DATE NAISSANCE" in df.columns and "DATE EFFET" in df.columns:
        age_r = (df["DATE EFFET"] - df["DATE NAISSANCE"]).dt.days / 365.25
        df["AGE SOUSCRIPTION"] = age_r.where(age_r.between(0,100)).round(0).astype("Int64")
        df["TRANCHE AGE"] = pd.cut(
            df["AGE SOUSCRIPTION"].astype(float),
            bins=[-np.inf,26,36,46,56,np.inf],
            labels=["[0-25]","[26-35]","[36-45]","[46-55]","[55+]"], right=False,
        ).astype(str).replace("nan","Indéfini")

    df["NB IMPAYES"] = (
        df.groupby("NUMERO POLICE")["NUMERO QUITTANCE"].transform("nunique").astype("Int16")
    )
    df["ANNEE"] = df["DEBUT PERIODE"].dt.year.astype("Int16")
    df["MOIS"]  = df["DEBUT PERIODE"].dt.to_period("M").astype(str)
    return df

# ══════════════════════════════════════════════════════════════════════════════
#  CHARGEMENT BASE LOCALE (Parquet)
# ══════════════════════════════════════════════════════════════════════════════
@st.cache_data(show_spinner=False)
def load_base() -> pd.DataFrame:
    if not base_exists():
        return pd.DataFrame()
    with st.spinner("📂 Chargement de la base locale en mémoire… Veuillez patienter."):
        df = pd.read_parquet(DATA_FILE)
        return enrich(df)

# ══════════════════════════════════════════════════════════════════════════════
#  HELPERS UI
# ══════════════════════════════════════════════════════════════════════════════
def fmt_fcfa(v: float) -> str:
    if pd.isna(v): return "–"
    if v >= 1e9:   return f"{v/1e9:,.2f} Mds F"
    if v >= 1e6:   return f"{v/1e6:,.1f} M F"
    return f"{v:,.0f} F".replace(",", " ")

def fmt_int(v) -> str:
    if pd.isna(v): return "0"
    return f"{int(v):,}".replace(",", " ")

def kpi_card(lbl, val, sub="", cls=""):
    return (f'<div class="kpi-card {cls}"><div class="kpi-label">{lbl}</div>'
            f'<div class="kpi-value {cls}">{val}</div><div class="kpi-sub">{sub}</div></div>')

def agg_tbl(df, grp, lbl=None):
    g = (df.groupby(grp)
           .agg(Nb_Quittances=("NUMERO QUITTANCE","nunique"),
                Nb_Polices   =("NUMERO POLICE",   "nunique"),
                Montant      =("MONTANT A PAYER", "sum"))
           .reset_index().sort_values("Montant",ascending=False))
    tot = g["Montant"].sum()
    g["% Total"]     = ((g["Montant"]/tot*100).round(1).astype(str)+" %") if tot > 0 else "0.0 %"
    g["Montant (F)"] = g["Montant"].apply(lambda x: f"{x:,.0f}".replace(",", " "))
    g.drop(columns="Montant", inplace=True)
    if lbl: g.rename(columns={grp: lbl}, inplace=True)
    return g

def opts(df, col):
    if col not in df.columns: return ["Tous"]
    return ["Tous"] + sorted(df[col].dropna().astype(str).unique().tolist())

def filt(df, col, val):
    if val == "Tous" or col not in df.columns: return df
    return df[df[col].astype(str) == str(val)]

def to_excel(df: pd.DataFrame, sheet: str = "Données") -> bytes:
    """
    Export Excel optimisé — openpyxl sans itération cellule par cellule.
    Style header uniquement (10× plus rapide que la version complète).
    """
    from openpyxl.styles import PatternFill, Font, Alignment
    from openpyxl.utils import get_column_letter

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name=sheet)
        ws = w.book[sheet]

        fill_hd = PatternFill("solid", fgColor="0B2559")
        font_hd = Font(color="C8A951", bold=True, size=10)
        font_bd = Font(size=9)
        align_c = Alignment(horizontal="center", vertical="center", wrap_text=True)
        align_l = Alignment(horizontal="left",   vertical="center")

        # ── Style entête (1 seule ligne) ──────────────────────────────────────
        for cell in ws[1]:
            cell.fill      = fill_hd
            cell.font      = font_hd
            cell.alignment = align_c

        # ── Largeur colonnes (échantillon 200 lignes, pas de formatage cellule) ─
        sample = df.head(200)
        for col_idx, col_name in enumerate(df.columns, 1):
            header_len = len(str(col_name))
            try:
                data_len = int(sample[col_name].astype(str).str.len().max())
            except Exception:
                data_len = header_len
            col_w = min(max(header_len, data_len) + 2, 40)
            ws.column_dimensions[get_column_letter(col_idx)].width = col_w

        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

    return buf.getvalue()

def dl_csv(df, fname, key):
    st.download_button(
        "⬇ CSV", df.to_csv(index=False, sep=";", encoding="utf-8-sig").encode("utf-8-sig"),
        fname, "text/csv", key=key, use_container_width=True,
    )

# ══════════════════════════════════════════════════════════════════════════════
#  HEADER
# ══════════════════════════════════════════════════════════════════════════════
st.markdown(f"""
<div class="header-banner">
  <div>
    <h1>💸 Suivi des Impayés</h1>
    <p>NSIA Vie Assurances · Direction des Études Réassurance et Actuariat</p>
  </div>
  <div><span class="header-badge">Mode LOCAL · SUN_COTEDIVOIRE</span></div>
</div>
""", unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
#  BARRE DE STATUT BASE LOCALE + BOUTONS GESTION
# ══════════════════════════════════════════════════════════════════════════════
meta = load_meta()

# ── Barre de statut ───────────────────────────────────────────────────────────
if base_exists():
    nb_lignes  = meta.get("nb_lignes", "?")
    last_update= meta.get("last_update", "inconnue")
    last_date  = meta.get("last_date_donnee", "?")
    first_date = meta.get("first_date", DATE_DEBUT_INIT)[:4]
    taille_mb  = round(os.path.getsize(DATA_FILE) / 1024 / 1024, 1)
    st.markdown(
        f'<div class="db-bar">'
        f'<span class="db-ok">✅ Base locale chargée</span>'
        f'<span class="db-info">📊 {fmt_int(nb_lignes)} lignes</span>'
        f'<span class="db-info">📅 Période : <b>{first_date} → {last_date}</b></span>'
        f'<span class="db-info">🕐 MAJ le : {last_update}</span>'
        f'<span class="db-info">💾 {taille_mb} Mo</span>'
        f'</div>',
        unsafe_allow_html=True,
    )
else:
    st.markdown(
        f'<div class="db-bar">'
        f'<span class="db-warn">⚠️ Aucune base locale.</span>'
        f'<span class="db-sub">Choisissez une année de départ ci-dessous et cliquez sur "📥 Télécharger la base".</span>'
        f'</div>',
        unsafe_allow_html=True,
    )

# ── Panneau de contrôle extraction ────────────────────────────────────────────
annee_courante = datetime.now().year

with st.expander("⚙️ Paramètres d'extraction — Choisir la période à charger", expanded=not base_exists()):

    st.markdown('<div class="fsec">Mode de sélection de la période</div>', unsafe_allow_html=True)

    mode_col, _, _ = st.columns([2, 2, 2])
    with mode_col:
        mode_extraction = st.radio(
            "Type de période",
            options=["📅 Depuis une année", "🗓️ Période personnalisée", "📦 Tout depuis 2015"],
            horizontal=True,
            key="mode_ext",
            label_visibility="collapsed",
        )

    # ── Mode 1 : depuis une année ─────────────────────────────────────────────
    if mode_extraction == "📅 Depuis une année":
        ANNEES_DISPO = list(range(2015, annee_courante + 1))
        c1, c2 = st.columns([2, 4])
        with c1:
            annee_debut = st.selectbox(
                "Année de départ",
                options=ANNEES_DISPO,
                index=0,
                format_func=lambda y: f"{y}  ({annee_courante - y + 1} an(s) de données)",
                key="sel_annee",
            )
        with c2:
            st.markdown(
                f"<div style='padding-top:1.8rem;font-size:.78rem;color:{GRAY};'>"
                f"⏱ Extraction du <b>01/01/{annee_debut}</b> au <b>{datetime.now().strftime('%d/%m/%Y')}</b> "
                f"→ environ <b>{(annee_courante - annee_debut + 1) * 2}</b> semestres</div>",
                unsafe_allow_html=True,
            )
        date_ext_debut = f"{annee_debut}0101"
        date_ext_fin   = datetime.now().strftime("%Y%m%d")

    # ── Mode 2 : période personnalisée ───────────────────────────────────────
    elif mode_extraction == "🗓️ Période personnalisée":
        c1, c2, c3 = st.columns([1.5, 1.5, 3])
        with c1:
            cal_deb = st.date_input(
                "Du (date de début)",
                value=date(annee_courante, 1, 1),
                min_value=date(2010, 1, 1),
                max_value=date.today(),
                key="cal_deb",
            )
        with c2:
            cal_fin = st.date_input(
                "Au (date de fin)",
                value=date.today(),
                min_value=date(2010, 1, 1),
                max_value=date.today(),
                key="cal_fin",
            )
        with c3:
            if cal_deb <= cal_fin:
                nb_mois = (cal_fin.year - cal_deb.year) * 12 + cal_fin.month - cal_deb.month + 1
                nb_sem  = max(1, round(nb_mois / 6))
                st.markdown(
                    f"<div style='padding-top:1.8rem;font-size:.78rem;color:{GRAY};'>"
                    f"⏱ Du <b>{cal_deb.strftime('%d/%m/%Y')}</b> au <b>{cal_fin.strftime('%d/%m/%Y')}</b> "
                    f"→ ~<b>{nb_mois}</b> mois, <b>{nb_sem}</b> semestre(s)</div>",
                    unsafe_allow_html=True,
                )
            else:
                st.error("⚠️ La date de fin doit être après la date de début.")
        date_ext_debut = cal_deb.strftime("%Y%m%d")
        date_ext_fin   = cal_fin.strftime("%Y%m%d")

    # ── Mode 3 : tout depuis 2015 ────────────────────────────────────────────
    else:
        st.markdown(
            f"<div style='font-size:.78rem;color:{GRAY};padding:.4rem 0;'>"
            f"📦 Extraction complète du <b>01/01/2015</b> au <b>{datetime.now().strftime('%d/%m/%Y')}</b> "
            f"→ environ <b>{(annee_courante - 2015 + 1) * 2}</b> semestres. "
            f"Durée estimée : <b>15-45 minutes</b> selon la vitesse réseau.</div>",
            unsafe_allow_html=True,
        )
        date_ext_debut = "20150101"
        date_ext_fin   = datetime.now().strftime("%Y%m%d")

    st.markdown("---")

    # ── Boutons ───────────────────────────────────────────────────────────────
    bc1, bc2, bc3 = st.columns([2, 1.5, 1.5])
    with bc1:
        btn_dl_base = st.button(
            "📥 Télécharger / Remplacer la base",
            help="Lance l'extraction sur la période choisie et remplace la base locale",
            use_container_width=True,
            type="primary",
        )
    with bc2:
        btn_maj = st.button(
            "🔄 Mettre à jour",
            help="Ajoute uniquement les nouvelles données depuis la dernière extraction",
            use_container_width=True,
        )
    with bc3:
        btn_reload = st.button(
            "↺ Rafraîchir l'affichage",
            help="Recharge le fichier Parquet en mémoire sans réextraction",
            use_container_width=True,
        )

if btn_reload:
    st.cache_data.clear()
    st.rerun()

# ══════════════════════════════════════════════════════════════════════════════
#  EXTRACTION INITIALE (année choisie → aujourd'hui)
# ══════════════════════════════════════════════════════════════════════════════
if btn_dl_base:
    chunks = generer_semestres(date_ext_debut, date_ext_fin)
    d_deb_lbl = f"{date_ext_debut[:4]}-{date_ext_debut[4:6]}-{date_ext_debut[6:]}"
    d_fin_lbl = f"{date_ext_fin[:4]}-{date_ext_fin[4:6]}-{date_ext_fin[6:]}"

    st.info(
        f"⏳ Extraction du **{d_deb_lbl}** au **{d_fin_lbl}** "
        f"en **{len(chunks)} semestre(s)**. Ne fermez pas l'application."
    )
    prog      = st.progress(0.0, text="Initialisation…")
    all_frames = []

    for i, (d1, d2) in enumerate(chunks, 1):
        label = f"Semestre {i}/{len(chunks)} : {d1[:4]}-{d1[4:6]} → {d2[:4]}-{d2[4:6]}"
        prog.progress(i / len(chunks), text=f"⬇ {label} — veuillez patienter…")
        try:
            chunk_df = fetch_chunk(d1, d2)
            if not chunk_df.empty:
                all_frames.append(chunk_df)
        except Exception as e:
            st.warning(f"⚠ Erreur {label} : {e} — semestre ignoré, on continue.")

    prog.progress(1.0, text="💾 Sauvegarde en cours…")

    if all_frames:
        df_full = pd.concat(all_frames, ignore_index=True)
        df_full.drop_duplicates(subset=["NUMERO QUITTANCE"], keep="last", inplace=True)
        df_full.to_parquet(DATA_FILE, index=False, compression="snappy")

        last_date = (
            df_full["DEBUT PERIODE"].dropna().astype(str).max()[:10]
            if "DEBUT PERIODE" in df_full.columns else today_str
        )
        save_meta({
            "nb_lignes":        len(df_full),
            "last_update":      datetime.now().strftime("%Y-%m-%d %H:%M"),
            "last_date_donnee": last_date,
            "first_date":       date_ext_debut,
        })
        st.cache_data.clear()
        st.success(f"✅ Base créée : **{len(df_full):,} lignes** extraites du {d_deb_lbl} au {last_date}")
        st.rerun()
    else:
        st.error("❌ Aucune donnée extraite. Vérifiez la connexion au serveur.")

# ══════════════════════════════════════════════════════════════════════════════
#  MISE À JOUR INCRÉMENTALE
# ══════════════════════════════════════════════════════════════════════════════
if btn_maj:
    if not base_exists():
        st.error("❌ Pas de base locale. Lancez d'abord '📥 Télécharger la base'.")
    else:
        meta = load_meta()
        last_date_str = meta.get("last_date_donnee", DATE_DEBUT_INIT).replace("-","")
        # Reprend depuis le lendemain de la dernière donnée
        last_dt  = datetime.strptime(last_date_str[:8], "%Y%m%d") + timedelta(days=1)
        today_dt = datetime.now()

        if last_dt.date() >= today_dt.date():
            st.info("✅ La base est déjà à jour.")
        else:
            d1 = last_dt.strftime("%Y%m%d")
            d2 = today_dt.strftime("%Y%m%d")
            chunks = generer_semestres(d1, d2)

            st.info(f"🔄 Mise à jour : {d1} → {d2} ({len(chunks)} chunk(s))")
            prog = st.progress(0.0, text="Démarrage…")
            new_frames = []

            for i, (c1, c2) in enumerate(chunks, 1):
                prog.progress(i / len(chunks), text=f"⬇ Chunk {i}/{len(chunks)} : {c1} → {c2}")
                try:
                    chunk_df = fetch_chunk(c1, c2)
                    if not chunk_df.empty:
                        new_frames.append(chunk_df)
                except Exception as e:
                    st.warning(f"⚠ Erreur chunk {c1}/{c2} : {e}")

            prog.progress(1.0, text="💾 Fusion et sauvegarde…")

            if new_frames:
                df_new  = pd.concat(new_frames, ignore_index=True)
                df_old  = pd.read_parquet(DATA_FILE)
                df_full = pd.concat([df_old, df_new], ignore_index=True)
                df_full.drop_duplicates(subset=["NUMERO QUITTANCE"], keep="last", inplace=True)
                df_full.to_parquet(DATA_FILE, index=False, compression="snappy")

                last_date = df_full["DEBUT PERIODE"].dropna().astype(str).max()[:10] if "DEBUT PERIODE" in df_full.columns else d2[:4]+"-"+d2[4:6]+"-"+d2[6:]
                save_meta({
                    "nb_lignes":        len(df_full),
                    "last_update":      datetime.now().strftime("%Y-%m-%d %H:%M"),
                    "last_date_donnee": last_date,
                    "first_date":       meta.get("first_date", DATE_DEBUT_INIT),
                })
                st.cache_data.clear()
                st.success(f"✅ +{len(df_new):,} nouvelles lignes · Base = {len(df_full):,} lignes au total")
                st.rerun()
            else:
                st.info("ℹ️ Aucune nouvelle donnée trouvée.")

# ══════════════════════════════════════════════════════════════════════════════
#  CHARGEMENT BASE + AFFICHAGE
# ══════════════════════════════════════════════════════════════════════════════
if not base_exists():
    st.stop()

df_base = load_base()
if df_base.empty:
    st.warning("⚠️ La base locale est vide ou corrompue. Relancez l'extraction.")
    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
#  SIDEBAR — filtres de période (Python pur, instantané)
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("### 🎛️ Filtres")
    st.markdown("---")

    # ── Sélection colonne de date ──────────────────────────────────────────
    st.markdown("**📅 Filtrer par date**")

    DATE_COLS_DISPO = {
        "Début de période":  "DEBUT PERIODE",
        "Fin de période":    "FIN PERIODE",
        "Date quittance":    "DATE QUITTANCE",
        "Date comptable":    "DATE COMPTABLE",
        "Date de création":  "DATE CREATION",
        "Date d'effet":      "DATE EFFET",
        "Date fin d'effet":  "DATE FIN EFFET",
        "Date résiliation":  "DATE RESILIATION",
        "Date naissance":    "DATE NAISSANCE",
    }
    # Ne proposer que les colonnes présentes dans la base chargée
    cols_dispo = {k: v for k, v in DATE_COLS_DISPO.items() if v in df_base.columns}

    col_date_label = st.selectbox("Colonne de date", list(cols_dispo.keys()), key="sb_col_date")
    col_date_sel   = cols_dispo[col_date_label]

    # Bornes dynamiques selon la colonne choisie
    _serie = df_base[col_date_sel].dropna()
    if len(_serie):
        dp_min = _serie.min().date()
        dp_max = _serie.max().date()
    else:
        dp_min = date(2015, 1, 1)
        dp_max = date.today()

    annee = datetime.now().year
    default_deb = max(dp_min, min(date(annee, 1, 1), dp_max))
    d_deb = st.date_input("Du", value=default_deb, min_value=dp_min, max_value=dp_max, key="d_deb")
    d_fin = st.date_input("Au", value=dp_max,      min_value=dp_min, max_value=dp_max, key="d_fin")

    st.markdown("---")
    with st.expander("🔍 Filtre agents"):
        agents_txt = st.text_area("Codes agents (un par ligne)", height=90,
                                  placeholder="Ex:\n012\n045")
    agents_list = [a.strip() for a in agents_txt.split("\n") if a.strip()]

    st.markdown("---")

    # ── Filtre DATE AFFECTATION — séparé et indépendant ──────────────────────
    st.markdown("**🏦 Date d'affectation (JAENCAP)**")
    st.caption("Date d'affectation de l'encaissement — indépendante des dates de quittance.")

    if "DATE AFFECTATION" in df_base.columns:
        _aff = df_base["DATE AFFECTATION"].dropna()
        if len(_aff):
            aff_min = _aff.min().date()
            aff_max = _aff.max().date()
        else:
            aff_min = date(2015, 1, 1)
            aff_max = date.today()
        annee_aff     = datetime.now().year
        default_aff_deb = max(aff_min, min(date(annee_aff, 1, 1), aff_max))
        aff_deb = st.date_input("Du", value=default_aff_deb,
                                min_value=aff_min, max_value=aff_max, key="aff_deb")
        aff_fin = st.date_input("Au", value=aff_max,
                                min_value=aff_min, max_value=aff_max, key="aff_fin")
        # Bouton reset pour désactiver le filtre
        if st.button("✖ Retirer ce filtre", key="reset_aff", use_container_width=True):
            st.session_state["aff_deb"] = aff_min
            st.session_state["aff_fin"] = aff_max
            st.rerun()
    else:
        st.caption("⚠ Colonne absente — relancez l'extraction.")
        aff_deb = None
        aff_fin = None

    st.markdown("---")
    st.markdown(f"""
    <div style="font-size:.72rem;color:#555;line-height:1.8;">
    🔒 <b>Règles figées :</b><br>
    ✅ INDIVIDUEL · INENC=' '<br>
    ✅ Hors polices 8xxx<br>
    ✅ Hors produits exclus<br>
    <br>
    <span style="color:#16a34a;font-weight:600;">⚡ Filtres 100% locaux — aucun appel SQL</span>
    </div>
    """, unsafe_allow_html=True)

# ══════════════════════════════════════════════════════════════════════════════
#  APPLICATION FILTRES (pandas en mémoire = instantané)
# ══════════════════════════════════════════════════════════════════════════════
dff = df_base.copy()

# Période — filtre sur la colonne choisie dans la sidebar
if col_date_sel in dff.columns:
    dff = dff[
        (dff[col_date_sel] >= pd.Timestamp(d_deb)) &
        (dff[col_date_sel] <= pd.Timestamp(d_fin))
    ]

# Agents
if agents_list:
    dff = dff[dff["CODE AGENT"].astype(str).isin(agents_list)]

# Date d'affectation — filtre indépendant (toujours actif si colonne présente)
if aff_deb is not None and aff_fin is not None and "DATE AFFECTATION" in dff.columns:
    dff = dff[
        (dff["DATE AFFECTATION"].isna()) |
        (
            (dff["DATE AFFECTATION"] >= pd.Timestamp(aff_deb)) &
            (dff["DATE AFFECTATION"] <= pd.Timestamp(aff_fin))
        )
    ]

# ── Filtres additionnels (expander) ──────────────────────────────────────────
with st.expander("🔍 Filtres additionnels (instantanés)", expanded=False):
    st.markdown('<div class="fsec">Mode / Motif / Durée / Ancienneté</div>', unsafe_allow_html=True)
    fc1,fc2,fc3,fc4 = st.columns(4)
    with fc1: f_mode  = st.selectbox("Mode encaissement",   opts(dff,"LIBELLE MODE"),       key="fm")
    with fc2: f_motif = st.selectbox("Motif prélèvement",   opts(dff,"LIBELLE MOTIF"),      key="fmo")
    with fc3: f_cls   = st.selectbox("Classe durée impayé", opts(dff,"CLASSE IMPAYES"),     key="fc")
    with fc4: f_anc   = st.selectbox("Ancienneté police",   opts(dff,"CLASSE ANCIENNETE"),  key="fa")

    st.markdown('<div class="fsec">Produit / Statut / Périodicité / Genre</div>', unsafe_allow_html=True)
    fc1,fc2,fc3,fc4 = st.columns(4)
    with fc1: f_prod  = st.selectbox("Code produit",    opts(dff,"CODE PRODUIT"),        key="fp")
    with fc2: f_stat  = st.selectbox("Statut police",   opts(dff,"STATUT POLICE"),       key="fs")
    with fc3: f_peri  = st.selectbox("Périodicité",     opts(dff,"LIBELLE PERIODICITE"), key="fpe")
    with fc4: f_genre = st.selectbox("Genre",           opts(dff,"GENRE"),              key="fg")

    st.markdown('<div class="fsec">Recherche libre</div>', unsafe_allow_html=True)
    f_srch = st.text_input("N° Police / NOM / Prénom", placeholder="Tapez…", key="fs2")

    st.markdown('<div class="fsec">⚡ Vues rapides</div>', unsafe_allow_html=True)
    VUES = ["Provision insuffisante","Compte clos","Compte inexistant",
            "Mobile / Numérique","≥ 12 impayés","Montant > 500 000 F"]
    vcols = st.columns(len(VUES))
    vue = None
    for i, v in enumerate(VUES):
        if vcols[i].button(v, key=f"v{i}", use_container_width=True): vue = v

# Appliquer
for col, val in [
    ("LIBELLE MODE",f_mode),("LIBELLE MOTIF",f_motif),("CLASSE IMPAYES",f_cls),
    ("CLASSE ANCIENNETE",f_anc),("CODE PRODUIT",f_prod),("STATUT POLICE",f_stat),
    ("LIBELLE PERIODICITE",f_peri),("GENRE",f_genre),
]:
    dff = filt(dff, col, val)

if f_srch:
    mask = pd.Series(False, index=dff.index)
    for c in ["NUMERO POLICE","NOM","PRENOMS","NOM COMPLET"]:
        if c in dff.columns:
            mask |= dff[c].astype(str).str.contains(f_srch, case=False, na=False)
    dff = dff[mask]

if vue == "Provision insuffisante":  dff = filt(dff,"LIBELLE MOTIF","Provision insuffisante")
elif vue == "Compte clos":           dff = filt(dff,"LIBELLE MOTIF","Compte clos")
elif vue == "Compte inexistant":     dff = filt(dff,"LIBELLE MOTIF","Compte inexistant")
elif vue == "Mobile / Numérique":    dff = filt(dff,"LIBELLE MODE", "Mobile / Numérique")
elif vue == "≥ 12 impayés" and "NB IMPAYES" in dff.columns:
    dff = dff[dff["NB IMPAYES"] >= 12]
elif vue == "Montant > 500 000 F":
    dff = dff[dff["MONTANT A PAYER"] > 500_000]

if dff.empty:
    st.warning("🔍 Aucun résultat — élargissez les filtres.")
    st.stop()

# ══════════════════════════════════════════════════════════════════════════════
#  KPIs
# ══════════════════════════════════════════════════════════════════════════════
total_imp  = dff["MONTANT A PAYER"].sum()
nb_pol     = dff["NUMERO POLICE"].nunique()
nb_quit    = dff["NUMERO QUITTANCE"].nunique()
nb_pol_tot = df_base["NUMERO POLICE"].nunique()
taux       = nb_pol / nb_pol_tot * 100 if nb_pol_tot else 0
moy_q      = total_imp / nb_quit if nb_quit else 0

st.markdown(
    '<div class="kpi-grid">'
    + kpi_card("Total Impayés (FCFA)",  fmt_fcfa(total_imp), "Cumul MONTANT A PAYER", "red")
    + kpi_card("Polices impayées",       fmt_int(nb_pol),     f"/ {fmt_int(nb_pol_tot)} base")
    + kpi_card("Quittances impayées",    fmt_int(nb_quit),    "uniques")
    + kpi_card("Taux impayé",            f"{taux:.1f} %",     "⚠ Critique" if taux>20 else "✓ OK",
               "red" if taux>20 else "green")
    + kpi_card("Montant moy / quitt.",   fmt_fcfa(moy_q),     "FCFA / quittance", "gold")
    + '</div>',
    unsafe_allow_html=True,
)

# ══════════════════════════════════════════════════════════════════════════════
#  ONGLETS
# ══════════════════════════════════════════════════════════════════════════════
t1, t2, t3, t4, t5, t6 = st.tabs([
    "📋 Extraction","🏦 Mode / Motif","📦 Produit / Statut",
    "⏱ Durée / Ancienneté","👤 Profil client","📈 Évolution",
])

fn_base_name = f"NSIA_Impayes_{d_deb:%Y%m%d}_{d_fin:%Y%m%d}"

# ─── T1 Extraction ─────────────────────────────────────────────────────────
with t1:
    try:
        st.markdown('<div class="section-title">Détail des quittances impayées</div>',unsafe_allow_html=True)

        # ── Colonnes organisées par groupes logiques ─────────────────────────
        COLS_PRIO = [
            # ── 1. Identification client ──────────────────────────────────────
            "NOM COMPLET", "NOM", "PRENOMS", "CIVILITE", "GENRE",
            "DATE NAISSANCE", "TRANCHE AGE", "AGE SOUSCRIPTION",
            "TELEPHONE", "TELEPHONE 2", "EMAIL", "ADRESSE POSTALE",

            # ── 2. Police ─────────────────────────────────────────────────────
            "NUMERO POLICE", "POLICE EXTERNE", "CODE PRODUIT",
            "NUMERO ASSURE",
            "CODE AGENT", "NOM AGENT", "CODE CLIENT",
            "DATE EFFET", "DATE FIN EFFET", "DATE RESILIATION",
            "ANCIENNETE ANS", "CLASSE ANCIENNETE",

            # ── 3. Quittance ──────────────────────────────────────────────────
            "NUMERO QUITTANCE",
            "DEBUT PERIODE", "FIN PERIODE",
            "DATE QUITTANCE", "DATE COMPTABLE", "DATE CREATION",
            "LIBELLE PERIODICITE",
            "NB IMPAYES", "DUREE IMPAYES", "CLASSE IMPAYES",

            # ── 4. Montants ───────────────────────────────────────────────────
            "MONTANT A PAYER", "MONTANT QUITTANCE",
            "SOLDE ENCAISSEMENT",

            # ── 5. Mode de paiement / Motif rejet ─────────────────────────────
            "LIBELLE MODE", "MODE ENCAISSEMENT",
            "LIBELLE MODE REGLEMENT", "MODE REGLEMENT",
            "LIBELLE MOTIF", "MOTIF PRELEVEMENT",

            # ── 6. Informations bancaires ─────────────────────────────────────
            "DATE AFFECTATION",
            "LIBELLE BANQUE", "LIBELLE AGENCE",
            "REF BANCAIRE ENC", "REF BANCAIRE BANQUE",
            "CODE IBAN", "LIBELLE ENCAISSEMENT DIRECT",
        ]

        # ── Colonnes attendues mais absentes du Parquet (extraction ancienne) ─
        COLS_BANQUE = ["LIBELLE BANQUE","LIBELLE AGENCE","REF BANCAIRE ENC",
                       "REF BANCAIRE BANQUE","CODE IBAN","LIBELLE ENCAISSEMENT DIRECT"]
        cols_manquantes = [c for c in COLS_BANQUE if c not in dff.columns]
        if cols_manquantes:
            st.warning(
                f"⚠️ Colonnes bancaires absentes de la base locale : "
                f"`{'`, `'.join(cols_manquantes)}`. "
                f"Ces colonnes ont été ajoutées récemment au SQL. "
                f"**Cliquez sur '📥 Télécharger la base' pour relancer l'extraction complète.**",
                icon="🏦",
            )
        sc = [c for c in COLS_PRIO if c in dff.columns]
        ds = dff[sc].copy()
        for dc in ["DATE NAISSANCE","DATE EFFET","DATE FIN EFFET","DATE RESILIATION",
                   "DEBUT PERIODE","FIN PERIODE","DATE QUITTANCE",
                   "DATE COMPTABLE","DATE CREATION","DATE AFFECTATION"]:
            if dc in ds.columns:
                ds[dc] = pd.to_datetime(ds[dc],errors="coerce").dt.strftime("%d/%m/%Y")

        # ── Définition des groupes de colonnes ────────────────────────────────
        GROUPES_COLS = {
            "👤 Client": [
                ("NOM COMPLET","Nom complet"),("NOM","Nom"),("PRENOMS","Prénoms"),
                ("CIVILITE","Civilité"),("GENRE","Genre"),("DATE NAISSANCE","Date naissance"),
                ("TRANCHE AGE","Tranche d'âge"),("TELEPHONE","Téléphone"),
                ("TELEPHONE 2","Téléphone 2"),("EMAIL","Email"),("ADRESSE POSTALE","Adresse"),
            ],
            "📋 Police": [
                ("NUMERO POLICE","N° Police"),("NUMERO ASSURE","N° Assuré"),
                ("CODE PRODUIT","Code produit"),("CODE AGENT","Code agent"),
                ("NOM AGENT","Nom agent"),("CODE CLIENT","Code client"),
                ("DATE EFFET","Date effet"),("DATE FIN EFFET","Fin effet"),
                ("DATE RESILIATION","Date résiliation"),
                ("ANCIENNETE ANS","Ancienneté (ans)"),("CLASSE ANCIENNETE","Classe ancienneté"),
            ],
            "🧾 Quittance": [
                ("NUMERO QUITTANCE","N° Quittance"),("DEBUT PERIODE","Début période"),
                ("FIN PERIODE","Fin période"),("DATE QUITTANCE","Date quittance"),
                ("DATE COMPTABLE","Date comptable"),("DATE CREATION","Date création"),
                ("LIBELLE PERIODICITE","Périodicité"),("NB IMPAYES","Nb impayés"),
                ("DUREE IMPAYES","Durée impayé (ans)"),("CLASSE IMPAYES","Classe durée"),
            ],
            "💰 Montants": [
                ("MONTANT A PAYER","Montant dû (FCFA)"),
                ("MONTANT QUITTANCE","Montant quittance"),
                ("SOLDE ENCAISSEMENT","Solde encaissement"),
            ],
            "🏦 Mode / Motif": [
                ("LIBELLE MODE","Mode encaissement"),("MODE ENCAISSEMENT","Code mode enc."),
                ("LIBELLE MODE REGLEMENT","Mode règlement"),("MODE REGLEMENT","Code mode règl."),
                ("LIBELLE MOTIF","Motif d'impayé"),("MOTIF PRELEVEMENT","Code motif"),
            ],
            "🏛️ Banque": [
                ("DATE AFFECTATION","Date d'affectation"),
                ("LIBELLE BANQUE","Banque"),("LIBELLE AGENCE","Agence"),
                ("REF BANCAIRE ENC","Réf. enc."),("REF BANCAIRE BANQUE","Réf. banque"),
                ("CODE IBAN","IBAN"),("LIBELLE ENCAISSEMENT DIRECT","Libellé enc. direct"),
            ],
        }
        # Colonnes non cochées par défaut (trop techniques)
        DESACTIVES_PAR_DEFAUT = {
            "POLICE EXTERNE","CODE CLIENT","MODE ENCAISSEMENT","MODE REGLEMENT",
            "MOTIF PRELEVEMENT","REF BANCAIRE ENC","REF BANCAIRE BANQUE","AGE SOUSCRIPTION",
        }
        all_avail = set(ds.columns)

        # ── Barre recherche + aperçu ──────────────────────────────────────────
        r1, r2 = st.columns([5, 1])
        with r1:
            cn, ca2 = st.columns(2)
            with cn:  s_nom = st.text_input("🔍 Nom", placeholder="Rechercher par nom…", key="s_nom")
            with ca2: s_agt = st.text_input("🔎 Agent", placeholder="Code ou nom…", key="s_agt")
        with r2:
            MAX_D = st.selectbox("Aperçu", [200,500,1000,2000,5000], index=1,
                                 label_visibility="collapsed", key="max_d")

        # ── Panneau "Extraction & Export" ─────────────────────────────────────
        with st.expander("📤  Extraction & Export — Choisir les colonnes et le format", expanded=False):

            # ── Sélection format ──────────────────────────────────────────────
            st.markdown(
                f"<div class='fsec'>FORMAT D'EXPORT</div>",
                unsafe_allow_html=True,
            )
            fmt_c1, fmt_c2, fmt_c3 = st.columns(3)
            with fmt_c1:
                if st.button("📊  Excel (.xlsx)", key="fmt_xl", use_container_width=True,
                             type="primary" if st.session_state.get("_fmt","excel")=="excel" else "secondary"):
                    st.session_state["_fmt"] = "excel"
            with fmt_c2:
                if st.button("📄  CSV (.csv)", key="fmt_csv", use_container_width=True,
                             type="primary" if st.session_state.get("_fmt","excel")=="csv" else "secondary"):
                    st.session_state["_fmt"] = "csv"
            with fmt_c3:
                st.markdown(
                    f"<div style='padding:.55rem 0;font-size:.75rem;color:{GRAY};'>"
                    f"Sélectionné : <b style='color:{NAVY};'>"
                    f"{'Excel' if st.session_state.get('_fmt','excel')=='excel' else 'CSV'}</b></div>",
                    unsafe_allow_html=True,
                )
            export_fmt = st.session_state.get("_fmt", "excel")

            # ── Sélection colonnes ────────────────────────────────────────────
            st.markdown("<div class='fsec'>COLONNES À INCLURE</div>", unsafe_allow_html=True)

            # Tout cocher / décocher
            ac1, ac2, _ = st.columns([1, 1, 4])
            with ac1:
                if st.button("✅ Tout cocher", key="chk_all", use_container_width=True):
                    for g, cols in GROUPES_COLS.items():
                        for cid, _ in cols:
                            if cid in all_avail:
                                st.session_state[f"xc_{cid}"] = True
            with ac2:
                if st.button("☐ Tout décocher", key="unchk_all", use_container_width=True):
                    for g, cols in GROUPES_COLS.items():
                        for cid, _ in cols:
                            st.session_state[f"xc_{cid}"] = False

            # Grille 2 colonnes
            grp_list = list(GROUPES_COLS.keys())
            half     = (len(grp_list) + 1) // 2
            gc1, gc2 = st.columns(2)
            for side, gcol in [(grp_list[:half], gc1), (grp_list[half:], gc2)]:
                with gcol:
                    for grp in side:
                        st.markdown(
                            f"<div style='font-size:.72rem;font-weight:700;color:{NAVY};"
                            f"margin:.7rem 0 .25rem;border-left:3px solid {GOLD};"
                            f"padding-left:7px;'>{grp}</div>",
                            unsafe_allow_html=True,
                        )
                        for cid, clbl in GROUPES_COLS[grp]:
                            if cid in all_avail:
                                default_v = cid not in DESACTIVES_PAR_DEFAUT
                                st.checkbox(
                                    clbl,
                                    value=st.session_state.get(f"xc_{cid}", default_v),
                                    key=f"xc_{cid}",
                                )
                            else:
                                st.markdown(
                                    f"<span style='font-size:.7rem;color:#ccc;'>"
                                    f"— {clbl} (non disponible)</span>",
                                    unsafe_allow_html=True,
                                )

            # Colonnes retenues
            cols_export = [
                cid
                for grp in GROUPES_COLS
                for cid, _ in GROUPES_COLS[grp]
                if cid in all_avail and st.session_state.get(
                    f"xc_{cid}", cid not in DESACTIVES_PAR_DEFAUT
                )
            ]
            ds_export = ds[cols_export] if cols_export else ds
            nb_exp    = len(ds_export)
            nb_col    = len(cols_export)

            st.markdown(
                f"<div style='font-size:.74rem;color:{GRAY};margin:.7rem 0 .4rem;padding:.5rem;"
                f"background:#f8f9fc;border-radius:6px;border:1px solid #e2e6f0;'>"
                f"📊 <b>{fmt_int(nb_exp)}</b> lignes · "
                f"<b>{nb_col}</b> colonne(s) sélectionnée(s) · "
                f"Format : <b>{'Excel (.xlsx)' if export_fmt=='excel' else 'CSV (.csv)'}</b>"
                f"</div>",
                unsafe_allow_html=True,
            )

            dl1, dl2 = st.columns(2)
            with dl1:
                st.download_button(
                    "⬇  Télécharger CSV",
                    data      = ds_export.to_csv(index=False, sep=";", encoding="utf-8-sig").encode("utf-8-sig"),
                    file_name = f"{fn_base_name}.csv",
                    mime      = "text/csv",
                    use_container_width=True,
                    key       = "dl_csv_exp",
                )
            with dl2:
                if st.button("⬇  Préparer Excel", key="btn_xl_exp",
                             use_container_width=True, type="primary"):
                    with st.spinner(f"⏳ Génération Excel — {fmt_int(nb_exp)} lignes × {nb_col} colonnes…"):
                        st.session_state["_xl_main"]  = to_excel(ds_export, "Impayés")
                        st.session_state["_xl_nb"]    = nb_exp

            if st.session_state.get("_xl_main"):
                st.success(f"✅ Fichier prêt — {fmt_int(st.session_state.get('_xl_nb', nb_exp))} lignes.")
                st.download_button(
                    "📥  Télécharger le fichier Excel",
                    data      = st.session_state["_xl_main"],
                    file_name = f"{fn_base_name}.xlsx",
                    mime      = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                    key       = "dl_xl_exp_ready",
                )

        # ── Aperçu tableau ────────────────────────────────────────────────────
        dv = ds.copy()
        if s_nom: dv = dv[dv["NOM COMPLET"].astype(str).str.contains(s_nom, case=False, na=False)]
        if s_agt:
            dv = dv[
                dv["CODE AGENT"].astype(str).str.contains(s_agt, case=False, na=False) |
                dv.get("NOM AGENT", pd.Series("", index=dv.index)).astype(str).str.contains(s_agt, case=False, na=False)
            ]

        st.markdown(
            f"<div style='font-size:.78rem;color:{GRAY};margin:.4rem 0 .3rem;'>"
            f"📊 <b>{fmt_int(len(dv))}</b> quittance(s) · "
            f"💰 Total : <b style='color:{RED}'>{fmt_fcfa(total_imp)}</b></div>",
            unsafe_allow_html=True,
        )
        st.dataframe(dv.head(MAX_D), use_container_width=True, height=420, hide_index=True)
        if len(dv) > MAX_D:
            st.info(
                f"⚠️ Aperçu limité à **{MAX_D} lignes**. "
                f"L'export contiendra les **{fmt_int(len(dv))} lignes complètes**."
            )
    except Exception as e:
        st.error(f"Erreur : {e}"); st.code(traceback.format_exc())

# ─── T2 Mode / Motif ───────────────────────────────────────────────────────
with t2:
    try:
        c1,c2 = st.columns(2)
        with c1:
            st.markdown('<div class="section-title">Mode d\'encaissement</div>',unsafe_allow_html=True)
            if "LIBELLE MODE" in dff.columns:
                g=agg_tbl(dff,"LIBELLE MODE","Mode")
                st.dataframe(g,use_container_width=True,hide_index=True)
                dl_csv(g,"mode.csv","dl_mode")
        with c2:
            st.markdown('<div class="section-title">Motif de rejet prélèvement</div>',unsafe_allow_html=True)
            if "LIBELLE MOTIF" in dff.columns:
                g=agg_tbl(dff,"LIBELLE MOTIF","Motif")
                st.dataframe(g,use_container_width=True,hide_index=True)
                dl_csv(g,"motif.csv","dl_motif")
        st.markdown('<div class="section-title">Top 20 agents</div>',unsafe_allow_html=True)
        if "CODE AGENT" in dff.columns:
            st.dataframe(agg_tbl(dff,"CODE AGENT","Agent").head(20),use_container_width=True,hide_index=True)
    except Exception as e: st.error(f"Erreur : {e}")

# ─── T3 Produit / Statut ───────────────────────────────────────────────────
with t3:
    try:
        c1,c2 = st.columns(2)
        with c1:
            st.markdown('<div class="section-title">Code produit</div>',unsafe_allow_html=True)
            if "CODE PRODUIT" in dff.columns:
                g=agg_tbl(dff,"CODE PRODUIT","Produit")
                st.dataframe(g,use_container_width=True,hide_index=True)
                dl_csv(g,"produit.csv","dl_prod")
        with c2:
            st.markdown('<div class="section-title">Statut police</div>',unsafe_allow_html=True)
            if "STATUT POLICE" in dff.columns:
                st.dataframe(agg_tbl(dff,"STATUT POLICE","Statut"),use_container_width=True,hide_index=True)
        c1,c2 = st.columns(2)
        with c1:
            st.markdown('<div class="section-title">Périodicité</div>',unsafe_allow_html=True)
            if "LIBELLE PERIODICITE" in dff.columns:
                st.dataframe(agg_tbl(dff,"LIBELLE PERIODICITE","Périodicité"),use_container_width=True,hide_index=True)
        with c2:
            st.markdown('<div class="section-title">Nb quittances / police</div>',unsafe_allow_html=True)
            if "NB IMPAYES" in dff.columns:
                st.dataframe(agg_tbl(dff,"NB IMPAYES","Nb impayés"),use_container_width=True,hide_index=True)
    except Exception as e: st.error(f"Erreur : {e}")

# ─── T4 Durée / Ancienneté ─────────────────────────────────────────────────
with t4:
    try:
        c1,c2 = st.columns(2)
        with c1:
            st.markdown('<div class="section-title">Classe durée impayé</div>',unsafe_allow_html=True)
            if "CLASSE IMPAYES" in dff.columns:
                ORD=["< 1 an","1-2 ans","2-3 ans","> 3 ans","Indéfini"]
                g=agg_tbl(dff,"CLASSE IMPAYES","Classe durée")
                g["Classe durée"]=pd.Categorical(g["Classe durée"],categories=ORD,ordered=True)
                st.dataframe(g.sort_values("Classe durée"),use_container_width=True,hide_index=True)
        with c2:
            st.markdown('<div class="section-title">Ancienneté police</div>',unsafe_allow_html=True)
            if "CLASSE ANCIENNETE" in dff.columns:
                ORD=["≤ 1 an","1-2 ans","2-3 ans","3-5 ans","5-10 ans","> 10 ans","Indéfini"]
                g=agg_tbl(dff,"CLASSE ANCIENNETE","Ancienneté")
                g["Ancienneté"]=pd.Categorical(g["Ancienneté"],categories=ORD,ordered=True)
                st.dataframe(g.sort_values("Ancienneté"),use_container_width=True,hide_index=True)
                dl_csv(g,"anciennete.csv","dl_anc")
    except Exception as e: st.error(f"Erreur : {e}")

# ─── T5 Profil client ──────────────────────────────────────────────────────
with t5:
    try:
        c1,c2 = st.columns(2)
        with c1:
            st.markdown('<div class="section-title">Genre</div>',unsafe_allow_html=True)
            if "GENRE" in dff.columns:
                st.dataframe(agg_tbl(dff,"GENRE","Genre"),use_container_width=True,hide_index=True)
        with c2:
            st.markdown('<div class="section-title">Tranche d\'âge (souscription)</div>',unsafe_allow_html=True)
            if "TRANCHE AGE" in dff.columns:
                ORD=["[0-25]","[26-35]","[36-45]","[46-55]","[55+]","Indéfini"]
                g=agg_tbl(dff,"TRANCHE AGE","Tranche âge")
                g["Tranche âge"]=pd.Categorical(g["Tranche âge"],categories=ORD,ordered=True)
                st.dataframe(g.sort_values("Tranche âge"),use_container_width=True,hide_index=True)
    except Exception as e: st.error(f"Erreur : {e}")

# ─── T6 Évolution ──────────────────────────────────────────────────────────
with t6:
    try:
        st.markdown('<div class="section-title">Évolution mensuelle</div>',unsafe_allow_html=True)
        if "MOIS" in dff.columns:
            g=agg_tbl(dff,"MOIS","Mois")
            st.dataframe(g.sort_values("Mois"),use_container_width=True,hide_index=True)
            dl_csv(g,"evolution_mensuelle.csv","dl_evo_m")
        st.markdown('<div class="section-title">Évolution annuelle</div>',unsafe_allow_html=True)
        if "ANNEE" in dff.columns:
            g=agg_tbl(dff,"ANNEE","Année")
            st.dataframe(g.sort_values("Année"),use_container_width=True,hide_index=True)
    except Exception as e: st.error(f"Erreur : {e}")

# ─── Pied de page ──────────────────────────────────────────────────────────
st.markdown(
    f"<div style='background:{NAVY};border-radius:8px;padding:10px 20px;margin-top:14px;"
    f"display:flex;justify-content:space-between;font-size:11px;color:rgba(255,255,255,.4)'>"
    f"<span>INDIVIDUEL · INENC=' ' · Hors 8xxx · Hors produits {sorted(PRODUITS_EXCLUS)}</span>"
    f"<span style='color:{GOLD}'>NSIA Vie Assurances — Direction Encaissement</span>"
    f"</div>",
    unsafe_allow_html=True,
)